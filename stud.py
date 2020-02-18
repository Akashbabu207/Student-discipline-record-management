 #@ Akash_babu_006
 #babushanthi207@gmail.com
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox, Frame, Label, Button, Entry      
from tkinter import font  as tkfont 
from tkinter import messagebox as tm
from PIL import ImageTk, Image
import sys
from collections import namedtuple
import csv
from openpyxl import Workbook
from tkinter import messagebox as tm
from openpyxl import load_workbook
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
#id and password of your mail id
login = 'mail id here'
password = 'password here'

global message1
wb = load_workbook("stud.xlsm")
ws2=wb["2cse"]
ws3=wb["3cse"]
ws4=wb["4cse"]
class hotel(tk.Tk):


     def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.winfo_toplevel().title("STUDENT DECIPLINE RECORD")

        self.title_font = tkfont.Font(family='Helvetica', size=18, weight="bold", slant="italic")
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        

        self.frames = {}
        for F in (firstpage,nextpage):
            page_name = F.__name__
            frame = F(parent=container, controller=self)
            self.frames[page_name] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame("firstpage")
        

     def show_frame(self, page_name):
        '''Show a frame for the given page name'''
        frame = self.frames[page_name]
        frame.tkraise()


     def msg(self):
    

        email_user =login
        email_password = password
        email_send = (txt4.get())

        subject = 'DISCIPLINE RECORD OF CSE'

        msg = MIMEMultipart()
        msg['From'] = email_user
        msg['To'] = email_send
        msg['Subject'] = subject

        body = 'THIS IS FROM CSE DEPARTMENT'
        msg.attach(MIMEText(body,'plain'))
        #ts = time.time()
        #date = datetime.datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
        filename="testexcel.xlsx"
        attachment  =open(filename,'rb')

        part = MIMEBase('application','octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',"attachment; filename= "+filename)

        msg.attach(part)
        text = msg.as_string()
        server = smtplib.SMTP('smtp.gmail.com',587)
        server.starttls()
        server.login(login,password)
        server.sendmail(login,email_send,text)
        server.quit()



     def view(self):
          a=(txt4.get())
          m=ws3.max_column
          if(a[0:2]=="27"):
               l=""
               ll="None"
               for i in range(2,100):
                    if((ws3.cell(row=i, column=3).value)==a):
                         for j in range(1,m+1):
                             c=ws3.cell(row=i,column=j+2)
                             
                             if(j+2<=5 ):
                                  l =l+" "+str(c.value)
                             else:
                                  ll=ll+" "+str(c.value)
                             lll=l+"       "+ll 
                             
                             #print(c.value,end=" ")
                         message1.configure(text= lll)
                         message1.config(font=('times', 15, 'italic'))
                         #message1.pack()
                         break

          
     
     def data(self):
          a=(txt.get())
          b=(txtt.get())
          row=[a,b]
          
          if(a[0:2]=="27"):
               
               for i in range(2,100):
                   if((ws3.cell(row=i, column=3).value)==a):
                       for j in range(5,100):
                           if((ws3.cell(row=i, column=j).value)==None):
                               ws3.cell(row=i,column=j).value=b
                               print("i")
                               print(ws3.cell(row=i, column=5).value)
                               if(j>=10):
                                    tm.message(text="ALert")
                               break
                       print()
          elif(a[0:2]=="28"):
               for i in range(2,100):
                   if((ws2.cell(row=i, column=3).value)==a):
                       for j in range(5,100):
                           if((ws2.cell(row=i, column=j).value)==None):
                               ws2.cell(row=i,column=j).value=b
                               print("j")
                               rint(ws2.cell(row=i, column=5).value)
                               break
                       print("balaji is prsent today")
          elif(a[0:2]=="29"):
               for i in range(2,100):
                   if((ws4.cell(row=i, column=3).value)==a):
                       for j in range(5,100):
                           if((ws4.cell(row=i, column=j).value)==None):
                               ws4.cell(row=i,column=j).value=b
                               print("k")
                               rint(ws4.cell(row=i, column=5).value)
                               break
                       print("balaji is prsent today")

          wb.save("testexcel.xlsx")
    
          

        
class firstpage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)

        #self.configure(background="red")

        load = Image.open("jjj.jpg")
        render = ImageTk.PhotoImage(load)
        img = Label(self, image=render)
        img.pack(side = "left", fill = "y", expand = "yes")
        img.image = render
        img.place(x=-475, y=-100)

       
        
        #self.heading=tk.Label(self,text="STUDENT DISCIPLINE RECORDS",bg="red",fg="black",width=30,height=1,font=("Algerian",45))
        #self.heading.place(x=70,y=20)


        self.bt2=tk.Label(self,text="ROLL NO           :",width=20,bg='black',fg='white',font=('arial',15,'bold'))
        self.bt2.place(x=500,y=300)

        global txt
        txt = tk.Entry(self,width=25,bg='#abb5c6' ,font=('times', 15, ' bold '))
        txt.pack(ipady=10)
        txt.place(x=800, y=300)
        
        #entry_1 = tk.Entry(self,width=30, bg='#abb5c6',font=("Verdana", 10))
        #entry_1.place(x=800, y=300)

        self.bt1=tk.Label(self,text="ISSUE                  :",width=20,bg='black',fg='white',font=('arial',15,'bold'))
        self.bt1.place(x=500,y=400)
        global txtt

        txtt = ttk.Combobox(self,width=39, values=["Shoe", "Shaving","Dress Code","Hair Cut","Mobile",])
        
        #txtt = tk.Entry(self,width=25,bg='#abb5c6' ,font=('times', 15, ' bold '))
        txtt.grid(column=0, row=1)
        txtt.pack(ipady=200)
        txtt.place(x=800, y=400)
        
        
        #entry_2 = tk.Entry(self,width=30,bg='#abb5c6',font=("Verdana", 10))
        #entry_2.place(x=800, y=400)

        self.bt3=tk.Button(self,text="SUBMIT",width=20,bg='black',fg='white',font=('arial',15,'bold'),command=lambda:controller.data())
        self.bt3.place(x=500,y=500)

        self.bt=tk.Button(self,text="<<NEXT PAGE>>",width=20,bg='black',fg='white',font=('arial',15,'bold'),command=lambda:controller.show_frame("nextpage"))
        self.bt.place(x=1100,y=700)


class nextpage(tk.Frame):
    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)
        self.controller = controller

        load = Image.open("jjj.jpg")
        render = ImageTk.PhotoImage(load)
        img = Label(self, image=render)
        img.pack(side = "left", fill = "y", expand = "yes")
        img.image = render
        img.place(x=-475, y=-100)
        
        #self.heading=tk.Label(self,text="Welurent",bg="white",fg="black",width=35,height=1,font=("times",35,'italic bold underline'))
        #elf.heading.place(x=275,y=10)

        self.lbl2=tk.Label(self,text="ROLL NO           :",width=20,bg='black',fg='white',font=('arial',15,'bold'))
        self.lbl2.place(x=500,y=300)

        global txt4
        txt4 = tk.Entry(self,width=25,bg='#abb5c6' ,font=('times', 15, ' bold '))
        txt4.pack(ipady=10)
        txt4.place(x=800, y=300)

        self.bt=tk.Button(self,text="<= previous page",width=20,bg='black',fg='white',font=('arial',15,'bold'),command=lambda:controller.show_frame("firstpage"))
        self.bt.place(x=1100,y=700)


        self.bt3=tk.Button(self,text="DETAIL",width=20,bg='black',fg='white',font=('arial',15,'bold'),command=lambda:controller.view())
        self.bt3.place(x=500,y=500)

        self.bt3=tk.Button(self,text="MAIL RECORDS",width=20,bg='black',fg='white',font=('arial',15,'bold'),command=lambda:controller.msg())
        self.bt3.place(x=500,y=600)

        global txt5
        txt5 = tk.Entry(self,width=25,bg='#abb5c6' ,font=('times', 15, ' bold '))
        txt5.pack(ipady=10)
        txt5.place(x=800, y=600)

        
        #self.bt=tk.Button(self,text="Next page =>",font=('arial',20,'bold'),command=lambda:controller.show_frame("nextpage2"))
        #self.bt.place(x=750,y=700)

        global message1
        message1 = tk.Label(self, text="",borderwidth=4, relief="solid" ,bg="white" ,fg="blue"  ,width=100  ,height=6, activebackground = "yellow" ,font=('times', 15, ' bold ')) 
        message1.place(x=300, y=350)




"""class nextpage2(tk.Frame):
    def __init__(self, parent, controller):

        tk.Frame.__init__(self, parent)
        self.controller = controller
        
        self.heading=tk.Label(self,text="Welurent",bg="white",fg="black",width=35,height=1,font=("times",35,'italic bold underline'))
        self.heading.place(x=275,y=10)

        self.bt=tk.Button(self,text="<= Previous page",font=('arial',20,'bold'),command=lambda:controller.show_frame("nextpage"))
        self.bt.place(x=700,y=700)"""




    

        
if __name__ == "__main__":
    app = hotel()
    app.geometry('1800x900')
    app.mainloop()
        

